import csv
from datetime import datetime, timedelta
from io import BytesIO
from collections import defaultdict

from django.shortcuts import get_object_or_404, render
from django.views import View
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from django.utils.timezone import localtime, now
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from accounts.models import Fingerprint, CustomUser
from .models import AttendanceEvent


class FingerprintAttendanceView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        finger_id = request.data.get("finger_id")
        template_data = request.data.get("template_data")
        device_id = request.data.get("device_id")

        if not finger_id or not template_data:
            return Response({"status": "error", "message": "finger_id و template_data الزامی است"}, status=400)

        fingerprint = get_object_or_404(Fingerprint, id=finger_id)
        fingerprint.template_data = template_data.encode('utf-8')
        fingerprint.save()

        user = fingerprint.user
        last_event = AttendanceEvent.objects.filter(user=user).order_by('-timestamp').first()
        current_time = now()

        if last_event and last_event.event_type == "in" and current_time - last_event.timestamp < timedelta(hours=24):
            event_type = "out"
        else:
            event_type = "in"

        AttendanceEvent.objects.create(
            user=user,
            event_type=event_type,
            verified_by="fingerprint",
            device_id=device_id,
            timestamp=current_time
        )

        return Response({
            "status": "success",
            "user": f"{user.first_name} {user.last_name}",
            "event_type": event_type,
            "timestamp": current_time.strftime("%Y-%m-%d %H:%M:%S")
        })

    def get(self, request):
        return Response({"status": "waiting"})


class AttendanceReportView(View):
    template_name = "report.html"
    table_template = "attendance_report_table.html"
    paginate_by = 25

    def get_users_list(self, request):
        users = CustomUser.objects.all().order_by('first_name', 'last_name')
        current_user = request.user
        user_role = getattr(current_user, 'roles', None)

        if user_role == 'user':
            users = users.filter(id=current_user.id)
        elif user_role == 'normal_admin' and hasattr(current_user, 'company') and current_user.company:
            users = users.filter(company=current_user.company)
        return users

    def calculate_work_hours(self, events):
        """
        محاسبه مجموع ساعت کاری روزانه برای هر کاربر
        خروجی: دیکشنری { (user_id, date) : timedelta(total_hours) }
        """
        work_hours = defaultdict(timedelta)
        events_by_user_date = defaultdict(list)

        for e in events:
            key = (e.user_id, e.timestamp.date())
            events_by_user_date[key].append(e)

        for key, ev_list in events_by_user_date.items():
            ev_list.sort(key=lambda x: x.timestamp)
            total = timedelta()
            in_time = None
            for ev in ev_list:
                if ev.event_type == 'in':
                    in_time = ev.timestamp
                elif ev.event_type == 'out' and in_time:
                    total += ev.timestamp - in_time
                    in_time = None
            work_hours[key] = total

        return work_hours

    def get(self, request):
        queryset = self.get_filtered_queryset(request)
        work_hours_dict = self.calculate_work_hours(queryset)

        export_type = request.GET.get("export", "").lower()
        if export_type in ["csv", "xlsx"]:
            return self.export_file(queryset, export_type, work_hours_dict)

        paginator = Paginator(queryset, self.paginate_by)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = {
            'page_obj': page_obj,
            'users': self.get_users_list(request),
            'work_hours_dict': work_hours_dict,  # اضافه شد
            'filter_values': {
                'user': request.GET.get('user', ''),
                'status': request.GET.get('status', ''),
                'start_date': request.GET.get('start_date', ''),
                'end_date': request.GET.get('end_date', ''),
                'report_url': request.get_full_path(),
            },
        }

        if request.headers.get('HX-Request'):
            return render(request, self.table_template, context)

        return render(request, self.template_name, context)

    def get_filtered_queryset(self, request):
        qs = AttendanceEvent.objects.select_related('user').order_by('-timestamp')
        current_user = request.user
        user_role = getattr(current_user, 'roles', None)

        if user_role == 'user':
            if not request.GET.get('user'):
                qs = qs.filter(user=current_user)
        elif user_role == 'normal_admin' and hasattr(current_user, 'company'):
            qs = qs.filter(user__company=current_user.company)

        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date and start_date.strip():
            try: qs = qs.filter(timestamp__date__gte=parse_date(start_date))
            except: pass
        if end_date and end_date.strip():
            try: qs = qs.filter(timestamp__date__lte=parse_date(end_date))
            except: pass

        nationality_number = request.GET.get('user')
        if nationality_number and nationality_number.strip():
            qs = qs.filter(user__nationality_number=nationality_number)

        status = request.GET.get('status')
        if status in ['complete', 'incomplete']:
            user_events = qs.values('user_id', 'event_type').distinct()
            users_with_in = {e['user_id'] for e in user_events if e['event_type'] == 'in'}
            users_with_out = {e['user_id'] for e in user_events if e['event_type'] == 'out'}
            if status == 'complete':
                allowed_users = users_with_in.intersection(users_with_out)
            else:
                allowed_users = users_with_in - users_with_out
            if allowed_users:
                qs = qs.filter(user_id__in=allowed_users)
            else:
                qs = qs.none()
        return qs

    def export_file(self, queryset, export_type, work_hours_dict):
        if export_type == "csv":
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="report_{datetime.now().strftime("%Y%m%d")}.csv"'
            response.write('\ufeff')
            writer = csv.writer(response)
            writer.writerow(['شماره ملی', 'نام کاربر', 'نوع رویداد', 'تاریخ و ساعت', 'ساعت کاری روز', 'تأیید کننده', 'دستگاه'])
            for event in queryset:
                user_name = f"{event.user.first_name} {event.user.last_name}"
                total_hours = work_hours_dict.get((event.user_id, event.timestamp.date()), timedelta())
                hours = total_hours.total_seconds() / 3600
                writer.writerow([
                    event.user.nationality_number or "-",
                    user_name,
                    "ورود" if event.event_type == "in" else "خروج",
                    localtime(event.timestamp).strftime("%Y/%m/%d %H:%M"),
                    f"{hours:.2f}",
                    event.verified_by or "-",
                    event.device_id or "-"
                ])
            return response

        elif export_type == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "گزارش تردد"
            headers = ['شماره ملی', 'نام کاربر', 'نوع رویداد', 'تاریخ و ساعت', 'ساعت کاری روز', 'تأیید کننده', 'شناسه دستگاه']
            ws.append(headers)
            header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill

            for event in queryset:
                user_name = f"{event.user.first_name} {event.user.last_name}"
                total_hours = work_hours_dict.get((event.user_id, event.timestamp.date()), timedelta())
                hours = total_hours.total_seconds() / 3600
                ws.append([
                    event.user.nationality_number or "-",
                    user_name,
                    "ورود" if event.event_type == "in" else "خروج",
                    localtime(event.timestamp).strftime("%Y/%m/%d %H:%M"),
                    f"{hours:.2f}",
                    event.verified_by or "-",
                    event.device_id or "-"
                ])

            ws.column_dimensions[get_column_letter(1)].width = 25
            for i in range(2, len(headers) + 1):
                ws.column_dimensions[get_column_letter(i)].width = 20

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            return response
